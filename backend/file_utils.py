"""
=============================================================================
FILE UTILITIES MODULE - ATOMIC FILE OPERATIONS
=============================================================================

This module provides utilities for safe file operations, with a focus on
atomic writes that protect against data corruption during unexpected
application shutdowns.

ATOMIC WRITE PATTERN:
---------------------
Instead of writing directly to the target file, we:
1. Create a temporary file in the SAME directory as the target
2. Write all content to the temporary file
3. Atomically rename the temporary file to the target filename

This ensures that if the process is interrupted at any point:
- Before step 3: Original file is completely untouched
- During step 3: Rename is atomic on same filesystem (OS-level guarantee)
- After step 3: Operation completed successfully

SECURITY & PRIVACY DESIGN DECISIONS:
-------------------------------------
1. SAME-DIRECTORY TEMP FILES: Temporary files are created in the same
   directory as the original file, NOT in system temp directories like
   /tmp or %TEMP%. This ensures:
   - Sensitive data never leaves the user's controlled location
   - Same filesystem guarantees atomic rename operation
   - Same permissions as the original file's directory
   - User can easily find and delete orphaned temp files

2. SECURE TEMP FILE NAMING: We use Python's tempfile module with a
   recognizable prefix (.groobi_tmp_) so orphaned files can be identified
   and cleaned up.

3. CLEANUP ON FAILURE: If any error occurs during writing, we attempt
   to clean up the temporary file before re-raising the exception.

USAGE:
------
    from file_utils import atomic_save_workbook
    
    wb = load_workbook(file_path)
    # ... modify workbook ...
    atomic_save_workbook(wb, file_path)

=============================================================================
"""

import os
import tempfile
import shutil
from typing import Optional
from contextlib import contextmanager

# Prefix for temporary files - makes orphaned files easy to identify
TEMP_FILE_PREFIX = ".groobi_tmp_"


def generate_temp_path(original_path: str) -> str:
    """
    Generates a temporary file path in the same directory as the original file.
    
    WHY SAME DIRECTORY:
    -------------------
    1. SECURITY: Sensitive data stays in user's controlled location
    2. ATOMICITY: Rename is only atomic on the same filesystem
    3. PERMISSIONS: Inherits parent directory permissions
    
    NAMING CONVENTION:
    ------------------
    The temp file is named: .groobi_tmp_<random>_<original_filename>
    
    Example:
        original: /path/to/data.xlsx
        temp:     /path/to/.groobi_tmp_abc123_data.xlsx
    
    The leading dot (.) hides the file on Unix systems, and the prefix
    makes it easy to identify and clean up orphaned temp files.
    
    Args:
        original_path: The path to the original file
        
    Returns:
        Path to a temporary file in the same directory
    """
    # Get the directory and filename of the original file
    directory = os.path.dirname(os.path.abspath(original_path))
    original_filename = os.path.basename(original_path)
    
    # Create a temporary file in the same directory
    # We use mkstemp to get a unique filename, then close and use the path
    fd, temp_path = tempfile.mkstemp(
        prefix=TEMP_FILE_PREFIX,
        suffix=f"_{original_filename}",
        dir=directory
    )
    
    # Close the file descriptor - we just needed the unique path
    os.close(fd)
    
    return temp_path


def atomic_rename(src_path: str, dst_path: str) -> None:
    """
    Atomically renames a file from src_path to dst_path.
    
    ATOMICITY GUARANTEE:
    --------------------
    On POSIX systems (Linux, macOS), os.rename() is atomic when both
    paths are on the same filesystem.
    
    On Windows, os.replace() provides the same atomic guarantee.
    
    We use os.replace() which works atomically on both platforms and
    will overwrite the destination if it exists.
    
    IMPORTANT:
    ----------
    This function assumes both paths are on the same filesystem.
    If they are on different filesystems, the operation may not be atomic.
    Our generate_temp_path() function ensures this by creating temp files
    in the same directory as the target.
    
    Args:
        src_path: Path to the source file (temporary file)
        dst_path: Path to the destination file (original file)
        
    Raises:
        OSError: If the rename operation fails
    """
    # os.replace() is atomic and works on both Windows and POSIX
    # It will overwrite dst_path if it exists
    os.replace(src_path, dst_path)


def cleanup_temp_file(temp_path: str) -> bool:
    """
    Safely removes a temporary file if it exists.
    
    This function is designed to be called in error handling scenarios
    where we need to clean up a partially-written temp file. It will
    not raise exceptions if the file doesn't exist or can't be deleted.
    
    Args:
        temp_path: Path to the temporary file to remove
        
    Returns:
        True if the file was deleted, False otherwise
    """
    try:
        if os.path.exists(temp_path):
            os.remove(temp_path)
            return True
    except OSError as e:
        # Log the error but don't raise - this is best-effort cleanup
        print(f"[file_utils] Warning: Could not clean up temp file {temp_path}: {e}")
    return False


def cleanup_orphaned_temp_files(directory: str) -> int:
    """
    Cleans up any orphaned temporary files in the specified directory.
    
    ORPHANED FILES:
    ---------------
    If the application crashes during a write operation, a temp file
    may be left behind. This function finds and removes such files.
    
    Files are identified by the TEMP_FILE_PREFIX (.groobi_tmp_).
    
    SECURITY NOTE:
    --------------
    This function only deletes files with our specific prefix, so it
    won't accidentally delete user files.
    
    Args:
        directory: The directory to scan for orphaned temp files
        
    Returns:
        The number of orphaned files that were deleted
    """
    deleted_count = 0
    
    try:
        for filename in os.listdir(directory):
            if filename.startswith(TEMP_FILE_PREFIX):
                file_path = os.path.join(directory, filename)
                if os.path.isfile(file_path):
                    if cleanup_temp_file(file_path):
                        deleted_count += 1
                        print(f"[file_utils] Cleaned up orphaned temp file: {filename}")
    except OSError as e:
        print(f"[file_utils] Warning: Error scanning directory {directory}: {e}")
    
    return deleted_count


@contextmanager
def atomic_write_context(target_path: str):
    """
    Context manager for atomic file writes.
    
    USAGE:
    ------
        with atomic_write_context("/path/to/file.xlsx") as temp_path:
            # Write to temp_path
            workbook.save(temp_path)
        # On successful exit, temp_path is atomically renamed to target_path
    
    ERROR HANDLING:
    ---------------
    If an exception occurs within the context:
    1. The temporary file is deleted (best-effort cleanup)
    2. The original file is left untouched
    3. The exception is re-raised
    
    Args:
        target_path: The final path where the file should end up
        
    Yields:
        The path to a temporary file where content should be written
    """
    temp_path = generate_temp_path(target_path)
    
    try:
        # Yield the temp path for the caller to write to
        yield temp_path
        
        # If we get here without exception, perform the atomic rename
        atomic_rename(temp_path, target_path)
        
    except Exception:
        # Clean up the temp file on any error
        cleanup_temp_file(temp_path)
        # Re-raise the original exception
        raise


def atomic_save_workbook(workbook, file_path: str) -> None:
    """
    Saves an openpyxl Workbook atomically.
    
    This is the main function you should use for saving Excel workbooks.
    It wraps the atomic write pattern in a simple, easy-to-use interface.
    
    HOW IT WORKS:
    -------------
    1. Creates a temporary file in the same directory as file_path
    2. Saves the workbook to the temporary file
    3. Atomically renames the temporary file to file_path
    
    If any step fails, the original file is left untouched.
    
    SECURITY:
    ---------
    - Temp file is created in the same directory (no data leakage)
    - Temp file has same permissions as directory
    - Temp file is cleaned up on failure
    
    Args:
        workbook: An openpyxl Workbook object
        file_path: The path where the workbook should be saved
        
    Raises:
        Exception: Any exception from workbook.save() or file operations
        
    Example:
        from openpyxl import load_workbook
        from file_utils import atomic_save_workbook
        
        wb = load_workbook("data.xlsx")
        ws = wb.active
        ws["A1"] = "Modified"
        atomic_save_workbook(wb, "data.xlsx")  # Safe atomic save
    """
    with atomic_write_context(file_path) as temp_path:
        workbook.save(temp_path)
