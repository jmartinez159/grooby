import unittest
from unittest.mock import patch, MagicMock
import pandas as pd
import math
import os
import openpyxl
from index import (
    normalize_value,
    get_target_sheets,
    get_common_columns_map,
    filter_noisy_columns,
    generate_signatures,
    highlight_rows,
    get_changed_rows,
    IGNORED_COLUMNS,
    NOISE_THRESHOLD
)

class TestIndex(unittest.TestCase):

    # --- normalize_value Tests ---
    def test_normalize_value_nan(self):
        self.assertEqual(normalize_value(float('nan')), "")
        self.assertEqual(normalize_value(None), "")

    def test_normalize_value_integer_float(self):
        self.assertEqual(normalize_value(10.0), "10")
        self.assertEqual(normalize_value(0.0), "0")

    def test_normalize_value_numbers(self):
        self.assertEqual(normalize_value(123), "123")
        self.assertEqual(normalize_value(12.34), "12.34")

    def test_normalize_value_string(self):
        self.assertEqual(normalize_value("  hello  "), "hello")
        self.assertEqual(normalize_value("test"), "test")

    # --- get_target_sheets Tests ---
    @patch('index.pd.ExcelFile')
    def test_get_target_sheets_success(self, mock_excel_file):
        mock_xls = MagicMock()
        mock_xls.sheet_names = ['Cover', 'Instructions', '1.1', '1.2']
        mock_excel_file.return_value = mock_xls
        
        result = get_target_sheets("dummy.xlsx")
        self.assertEqual(result, ('1.1', '1.2'))

    @patch('index.pd.ExcelFile')
    def test_get_target_sheets_too_few(self, mock_excel_file):
        mock_xls = MagicMock()
        mock_xls.sheet_names = ['Cover', '1.1'] # Only 1 matching
        mock_excel_file.return_value = mock_xls
        
        result = get_target_sheets("dummy.xlsx")
        self.assertEqual(result, (None, None))

    @patch('index.pd.ExcelFile')
    def test_get_target_sheets_exception(self, mock_excel_file):
        mock_excel_file.side_effect = Exception("File error")
        result = get_target_sheets("dummy.xlsx")
        self.assertEqual(result, (None, None))

    # --- get_common_columns_map Tests ---
    def test_get_common_columns_map(self):
        # Setup DataFrames
        data_prev = {
            'A': ['Header1', 'Data'],
            'B': ['Header2', 'Data'],
            'C': ['LOT #', 'Data'] # Should be ignored
        }
        df_prev = pd.DataFrame(data_prev)
        
        # DataFrame constructor from dict uses keys as columns.
        # So columns are A, B, C.
        # But data is:
        # A: Header1, Data
        # B: Header2, Data
        # C: LOT #, Data
        
        # IMPORTANT: index.py uses df.iloc[0] to get headers.
        # row `0` is ['Header1', 'Header2', 'LOT #']
        # row `1` is ['Data', 'Data', 'Data']
        
        # df_prev columns are A, B, C.
        # row 0 index is 0.
        # map logic iterates over row_series.items(): (col_name, val).
        # col_name is A (or index 0? No, dataframe iteration).
        # items() on series returns (index, value).
        # For a row from a DF, the index is the Column Name.
        
        # Wait. index.py logic:
        # row_series = df.iloc[0]
        # for col_index, val in row_series.items():
        #     ...
        # In Pandas, if you take a row from a DataFrame, it returns a Series where the index is the DataFrame's columns.
        # In `index.py`, the DFs are created with:
        # pd.read_excel(..., header=None, skiprows=1).
        # So columns are Int64Index([0, 1, 2, ...]).
        
        # In my test `df_prev = pd.DataFrame(data_prev)`, columns are Strings 'A', 'B', 'C'.
        # So `col_index` in the loop will be 'A', 'B', 'C'.
        # But `index.py` expects `col_index` to be an integer column index?
        # `value_map[clean_val].append(col_index)` -> `results[val] = {..., prev_indices: ...}`
        
        # Later, `filter_noisy_columns`:
        # `data_prev.iloc[:min_len, idx_prev].apply(...)`
        # `idx_prev` comes from `col_index`.
        # If `idx_prev` is 'A', `iloc` with a string might fail if specific column access semantics differ or work (pandas allows label access in loc, but iloc strictly requires integers).
        # `iloc` requires integers.
        # So `col_index` MUST be an integer.
        
        # ERROR FOUND: `index.py` relies on `header=None` which gives integer column names. 
        # My test creates DFs with string columns.
        # FIX: Create DFs with integer columns in tests.
        
        df_prev = pd.DataFrame([
            ['Header1', 'Header2', 'LOT #'],
            ['Data', 'Data', 'Data']
        ])
        # Default columns are 0, 1, 2.
        
        df_curr = pd.DataFrame([
            ['Header2', 'Header1', 'New'],
            ['Data', 'Data', 'Data']
        ])
        
        result = get_common_columns_map(df_prev, df_curr)
        
        expected = {
            'Header1': {'prev_indices': [0], 'curr_indices': [1]},
            'Header2': {'prev_indices': [1], 'curr_indices': [0]}
        }
        self.assertEqual(result, expected)

    def test_get_common_columns_duplicates(self):
        # Case where a header appears multiple times
        data_prev = pd.DataFrame([['H1', 'H1', 'H2']])
        data_curr = pd.DataFrame([['H1', 'H2', 'H2']])
        
        result = get_common_columns_map(data_prev, data_curr)
        
        # H1 in prev: [0, 1], in curr: [0]
        # H2 in prev: [2], in curr: [1, 2]
        self.assertEqual(result['H1']['prev_indices'], [0, 1])
        self.assertEqual(result['H1']['curr_indices'], [0])
        self.assertEqual(result['H2']['prev_indices'], [2])
        self.assertEqual(result['H2']['curr_indices'], [1, 2])
        
    # --- filter_noisy_columns Tests ---
    def test_filter_noisy_columns_no_noise(self):
        # Matches: Col A (IDX 0) -> Col A (IDX 0)
        matches = {'ColA': {'prev_indices': [0], 'curr_indices': [0]}}
        
        # Row 0 is header (skipped by iloc[1:])
        # Row 1.. are data
        df_prev = pd.DataFrame([['ColA'], ['1'], ['2'], ['3']])
        df_curr = pd.DataFrame([['ColA'], ['1'], ['2'], ['3']])
        
        result = filter_noisy_columns(df_prev, df_curr, matches)
        self.assertIn('ColA', result)

    def test_filter_noisy_columns_noisy(self):
        matches = {'ColA': {'prev_indices': [0], 'curr_indices': [0]}}
        
        # 4 data rows. 3 changes -> 75% change > 50% threshold
        df_prev = pd.DataFrame([['ColA'], ['1'], ['2'], ['3'], ['4']])
        df_curr = pd.DataFrame([['ColA'], ['1'], ['X'], ['Y'], ['Z']])
        
        result = filter_noisy_columns(df_prev, df_curr, matches)
        self.assertNotIn('ColA', result)
        
    def test_filter_noisy_columns_empty_data(self):
        matches = {'ColA': {'prev_indices': [0], 'curr_indices': [0]}}
        df_prev = pd.DataFrame([['ColA']]) # No data rows
        df_curr = pd.DataFrame([['ColA']])
        
        result = filter_noisy_columns(df_prev, df_curr, matches)
        # Should return matches as is
        self.assertEqual(result, matches)

    def test_filter_noisy_columns_mismatched_length(self):
        matches = {'ColA': {'prev_indices': [0], 'curr_indices': [0]}}
        df_prev = pd.DataFrame([['ColA'], ['1'], ['2']])
        df_curr = pd.DataFrame([['ColA'], ['1']])
        
        # Min len is 1 (row '1'). No changes.
        result = filter_noisy_columns(df_prev, df_curr, matches)
        self.assertIn('ColA', result)

    # --- generate_signatures Tests ---
    def test_generate_signatures(self):
        matches = {
            'A': {'prev_indices': [0], 'curr_indices': [1]},
            'B': {'prev_indices': [1], 'curr_indices': [0]}
        }
        # Sorted keys: A, B
        # extract_cols for prev: [0, 1]
        
        df = pd.DataFrame([
            ['HeaderA', 'HeaderB'], # Row 0 (excluded)
            ['ValA1', 'ValB1'],     # Row 1 -> Sig: ValA1-ValB1
            ['ValA2', 'ValB2']      # Row 2 -> Sig: ValA2-ValB2
        ])
        
        sigs = generate_signatures(df, matches, 'prev_indices')
        self.assertEqual(sigs, ['ValA1-ValB1', 'ValA2-ValB2'])

    def test_generate_signatures_out_of_bounds(self):
        # Case where index points outside row (should be handled gracefully with "")
        matches = {'A': {'prev_indices': [5], 'curr_indices': [5]}} 
        df = pd.DataFrame([['H'], ['V']])
        
        sigs = generate_signatures(df, matches, 'prev_indices')
        self.assertEqual(sigs, ['']) # Empty string for out of bounds

    # --- highlight_rows Tests ---
    @patch('index.atomic_save_workbook')
    @patch('index.load_workbook')
    @patch('index.PatternFill')
    def test_highlight_rows_success(self, mock_fill, mock_load_wb, mock_atomic_save):
        mock_wb = MagicMock()
        mock_ws = MagicMock()
        mock_cell = MagicMock()
        
        mock_wb.sheetnames = ['Sheet1']
        mock_wb.__getitem__.return_value = mock_ws
        
        # ws[row] should be iterable (cells)
        mock_ws.__getitem__.return_value = [mock_cell, mock_cell]
        
        mock_load_wb.return_value = mock_wb
        
        indices = [0, 2] 
        highlight_rows("test.xlsx", "Sheet1", indices)
        
        # Check workbook loaded
        mock_load_wb.assert_called_with("test.xlsx")
        # Check atomic save called (not wb.save directly)
        mock_atomic_save.assert_called_with(mock_wb, "test.xlsx")

    @patch('index.load_workbook')
    def test_highlight_rows_sheet_not_found(self, mock_load_wb):
        mock_wb = MagicMock()
        mock_wb.sheetnames = ['Other']
        mock_load_wb.return_value = mock_wb
        
        highlight_rows("test.xlsx", "Target", [1])
        mock_wb.save.assert_not_called()

    @patch('index.load_workbook')
    def test_highlight_rows_exception(self, mock_load_wb):
        mock_load_wb.side_effect = Exception("Disk error")
        # Should catch and print, not raise
        highlight_rows("test.xlsx", "Sheet1", [1])
        # Pass implies success

    # --- get_changed_rows (Integration Logic) Tests ---
    
    @patch('index.get_target_sheets')
    @patch('index.pd.read_excel')
    @patch('index.highlight_rows')
    @patch('os.path.exists')
    def test_get_changed_rows_full_flow(self, mock_exists, mock_highlight, mock_read_excel, mock_sheets):
        mock_exists.return_value = True
        mock_sheets.return_value = ('v1', 'v2')
        
        # Setup DataFrames
        # v1: Header A, Row 1, Row 2, Row 3
        # v2: Header A, Row 1 (Changed), Row 2 (Same), Row 3 (Same)
        # 3 data rows. 1 change. 33% change < 50% threshold.
        df_prev = pd.DataFrame([['A'], ['original'], ['keep'], ['keep']])
        df_curr = pd.DataFrame([['A'], ['changed'], ['keep'], ['keep']])
        
        mock_read_excel.side_effect = [df_prev, df_curr]
        
        # Run
        changed = get_changed_rows("test.xlsx")
        
        # Expect changes
        # Row 0 in df_curr data (iloc[1:]) is 'changed'.
        # normalize_value('original') vs 'changed'. -> Changed.
        # Index in curr data list is 0. +1 => return index 1 relative to df_curr
        
        self.assertIsNotNone(changed)
        self.assertFalse(changed.empty)
        # Verify highlighting called
        mock_highlight.assert_called()

    @patch('os.path.exists')
    def test_get_changed_rows_no_file(self, mock_exists):
        mock_exists.return_value = False
        res = get_changed_rows("missing.xlsx")
        self.assertIsNone(res)

    @patch('os.path.exists')
    @patch('index.get_target_sheets')
    def test_get_changed_rows_no_sheets(self, mock_sheets, mock_exists):
        mock_exists.return_value = True
        mock_sheets.return_value = (None, None)
        res = get_changed_rows("test.xlsx")
        self.assertIsNone(res)

    @patch('os.path.exists')
    @patch('index.get_target_sheets')
    @patch('index.pd.read_excel')
    def test_get_changed_rows_no_common_cols(self, mock_read, mock_sheets, mock_exists):
        mock_exists.return_value = True
        mock_sheets.return_value = ('v1', 'v2')
        # Disjoint headers
        mock_read.side_effect = [
            pd.DataFrame([['A'], ['1']]),
            pd.DataFrame([['B'], ['1']])
        ]
        res = get_changed_rows("test.xlsx")
        self.assertIsNone(res)
        
    @patch('os.path.exists')
    @patch('index.get_target_sheets')
    @patch('index.pd.read_excel')
    def test_get_changed_rows_all_noisy(self, mock_read, mock_sheets, mock_exists):
        mock_exists.return_value = True
        mock_sheets.return_value = ('v1', 'v2')
        
        # Make it noisy
        # 1 col, 2 rows. 2 changes. 100% change.
        df_prev = pd.DataFrame([['A'], ['1'], ['2']])
        df_curr = pd.DataFrame([['A'], ['X'], ['Y']])
        
        mock_read.side_effect = [df_prev, df_curr]
        
        res = get_changed_rows("test.xlsx")
        self.assertIsNone(res) # Should filter out and return None

    @patch('os.path.exists')
    @patch('index.get_target_sheets')
    @patch('index.pd.read_excel')
    def test_get_changed_rows_no_changes(self, mock_read, mock_sheets, mock_exists):
        mock_exists.return_value = True
        mock_sheets.return_value = ('v1', 'v2')
        
        df_prev = pd.DataFrame([['A'], ['1']])
        df_curr = pd.DataFrame([['A'], ['1']])
        
        mock_read.side_effect = [df_prev, df_curr]
        
        res = get_changed_rows("test.xlsx")
        self.assertTrue(res.empty)

if __name__ == '__main__':
    unittest.main()
