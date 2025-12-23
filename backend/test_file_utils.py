"""
=============================================================================
UNIT TESTS FOR FILE UTILITIES MODULE
=============================================================================

This test module provides comprehensive coverage for the atomic file
operations in file_utils.py.

TEST STRATEGY:
--------------
1. Test each function in isolation
2. Test the atomic write pattern end-to-end
3. Test error handling and cleanup scenarios
4. Test security properties (same-directory temp files)

FIXTURES:
---------
We use pytest's tmp_path fixture for creating temporary test directories.
This ensures test isolation and automatic cleanup.

=============================================================================
"""

import os
import pytest
from unittest.mock import Mock, patch, MagicMock
from openpyxl import Workbook

# Import the module under test
from file_utils import (
    TEMP_FILE_PREFIX,
    generate_temp_path,
    atomic_rename,
    cleanup_temp_file,
    cleanup_orphaned_temp_files,
    atomic_write_context,
    atomic_save_workbook,
)


# =============================================================================
# TEST: generate_temp_path()
# =============================================================================

class TestGenerateTempPath:
    """Tests for the generate_temp_path function."""
    
    def test_temp_file_created_in_same_directory(self, tmp_path):
        """
        SECURITY TEST: Temp file must be in the same directory as original.
        
        This is critical for:
        1. Atomic rename guarantee (same filesystem)
        2. Data privacy (no leakage to system temp)
        """
        # Arrange: Create a test file path
        original_file = tmp_path / "data.xlsx"
        original_file.touch()
        
        # Act: Generate temp path
        temp_path = generate_temp_path(str(original_file))
        
        # Assert: Temp file should be in same directory
        assert os.path.dirname(temp_path) == str(tmp_path)
        
        # Cleanup
        if os.path.exists(temp_path):
            os.remove(temp_path)
    
    def test_temp_file_has_correct_prefix(self, tmp_path):
        """Temp file should have our identifiable prefix for cleanup."""
        # Arrange
        original_file = tmp_path / "data.xlsx"
        original_file.touch()
        
        # Act
        temp_path = generate_temp_path(str(original_file))
        temp_filename = os.path.basename(temp_path)
        
        # Assert
        assert temp_filename.startswith(TEMP_FILE_PREFIX)
        
        # Cleanup
        if os.path.exists(temp_path):
            os.remove(temp_path)
    
    def test_temp_file_contains_original_filename(self, tmp_path):
        """Temp file should contain original filename for identification."""
        # Arrange
        original_file = tmp_path / "my_data.xlsx"
        original_file.touch()
        
        # Act
        temp_path = generate_temp_path(str(original_file))
        temp_filename = os.path.basename(temp_path)
        
        # Assert
        assert "my_data.xlsx" in temp_filename
        
        # Cleanup
        if os.path.exists(temp_path):
            os.remove(temp_path)
    
    def test_temp_file_is_unique(self, tmp_path):
        """Each call should generate a unique temp path."""
        # Arrange
        original_file = tmp_path / "data.xlsx"
        original_file.touch()
        
        # Act
        temp_path_1 = generate_temp_path(str(original_file))
        temp_path_2 = generate_temp_path(str(original_file))
        
        # Assert
        assert temp_path_1 != temp_path_2
        
        # Cleanup
        for path in [temp_path_1, temp_path_2]:
            if os.path.exists(path):
                os.remove(path)


# =============================================================================
# TEST: atomic_rename()
# =============================================================================

class TestAtomicRename:
    """Tests for the atomic_rename function."""
    
    def test_rename_replaces_destination(self, tmp_path):
        """Atomic rename should replace the destination file."""
        # Arrange
        src_file = tmp_path / "source.txt"
        dst_file = tmp_path / "destination.txt"
        
        src_file.write_text("new content")
        dst_file.write_text("old content")
        
        # Act
        atomic_rename(str(src_file), str(dst_file))
        
        # Assert
        assert not src_file.exists()  # Source should be gone
        assert dst_file.exists()  # Destination should exist
        assert dst_file.read_text() == "new content"  # Content should be new
    
    def test_rename_creates_destination_if_not_exists(self, tmp_path):
        """Atomic rename should work even if destination doesn't exist."""
        # Arrange
        src_file = tmp_path / "source.txt"
        dst_file = tmp_path / "new_file.txt"
        
        src_file.write_text("content")
        
        # Act
        atomic_rename(str(src_file), str(dst_file))
        
        # Assert
        assert not src_file.exists()
        assert dst_file.exists()
        assert dst_file.read_text() == "content"


# =============================================================================
# TEST: cleanup_temp_file()
# =============================================================================

class TestCleanupTempFile:
    """Tests for the cleanup_temp_file function."""
    
    def test_deletes_existing_file(self, tmp_path):
        """Should delete the file if it exists."""
        # Arrange
        temp_file = tmp_path / "temp.txt"
        temp_file.write_text("content")
        
        # Act
        result = cleanup_temp_file(str(temp_file))
        
        # Assert
        assert result is True
        assert not temp_file.exists()
    
    def test_returns_false_for_nonexistent_file(self, tmp_path):
        """Should return False if file doesn't exist."""
        # Arrange
        nonexistent = tmp_path / "nonexistent.txt"
        
        # Act
        result = cleanup_temp_file(str(nonexistent))
        
        # Assert
        assert result is False
    
    def test_handles_permission_error_gracefully(self, tmp_path):
        """Should not raise exception on permission error."""
        # Arrange
        temp_file = tmp_path / "temp.txt"
        temp_file.write_text("content")
        
        # Act & Assert: Mock os.remove to raise PermissionError
        with patch('file_utils.os.remove', side_effect=PermissionError("Access denied")):
            with patch('file_utils.os.path.exists', return_value=True):
                result = cleanup_temp_file(str(temp_file))
        
        # Should return False and not raise
        assert result is False


# =============================================================================
# TEST: cleanup_orphaned_temp_files()
# =============================================================================

class TestCleanupOrphanedTempFiles:
    """Tests for the cleanup_orphaned_temp_files function."""
    
    def test_deletes_files_with_temp_prefix(self, tmp_path):
        """Should delete files that start with our temp prefix."""
        # Arrange: Create orphaned temp files
        orphan1 = tmp_path / f"{TEMP_FILE_PREFIX}abc_data.xlsx"
        orphan2 = tmp_path / f"{TEMP_FILE_PREFIX}xyz_other.xlsx"
        orphan1.write_text("content")
        orphan2.write_text("content")
        
        # Act
        deleted_count = cleanup_orphaned_temp_files(str(tmp_path))
        
        # Assert
        assert deleted_count == 2
        assert not orphan1.exists()
        assert not orphan2.exists()
    
    def test_does_not_delete_regular_files(self, tmp_path):
        """Should NOT delete files without our temp prefix."""
        # Arrange
        regular_file = tmp_path / "data.xlsx"
        other_temp = tmp_path / ".some_other_temp_file"
        regular_file.write_text("important data")
        other_temp.write_text("other temp")
        
        # Act
        deleted_count = cleanup_orphaned_temp_files(str(tmp_path))
        
        # Assert
        assert deleted_count == 0
        assert regular_file.exists()
        assert other_temp.exists()
    
    def test_handles_empty_directory(self, tmp_path):
        """Should handle empty directory gracefully."""
        # Act
        deleted_count = cleanup_orphaned_temp_files(str(tmp_path))
        
        # Assert
        assert deleted_count == 0


# =============================================================================
# TEST: atomic_write_context()
# =============================================================================

class TestAtomicWriteContext:
    """Tests for the atomic_write_context context manager."""
    
    def test_successful_write_replaces_original(self, tmp_path):
        """On success, temp file should replace the original."""
        # Arrange
        target_file = tmp_path / "data.txt"
        target_file.write_text("original content")
        
        # Act
        with atomic_write_context(str(target_file)) as temp_path:
            with open(temp_path, 'w') as f:
                f.write("new content")
        
        # Assert
        assert target_file.read_text() == "new content"
        assert not os.path.exists(temp_path)  # Temp was renamed
    
    def test_failed_write_preserves_original(self, tmp_path):
        """On failure, original file should be untouched."""
        # Arrange
        target_file = tmp_path / "data.txt"
        target_file.write_text("original content")
        
        # Act & Assert
        with pytest.raises(ValueError):
            with atomic_write_context(str(target_file)) as temp_path:
                with open(temp_path, 'w') as f:
                    f.write("partial content")
                raise ValueError("Simulated error")
        
        # Original should be untouched
        assert target_file.read_text() == "original content"
        # Temp file should be cleaned up
        assert not os.path.exists(temp_path)
    
    def test_temp_file_in_same_directory(self, tmp_path):
        """Temp file should be created in same directory as target."""
        # Arrange
        target_file = tmp_path / "data.txt"
        target_file.write_text("content")
        
        # Act
        with atomic_write_context(str(target_file)) as temp_path:
            # Assert: Check temp is in same directory
            assert os.path.dirname(temp_path) == str(tmp_path)
            
            # Write something so the context completes
            with open(temp_path, 'w') as f:
                f.write("content")


# =============================================================================
# TEST: atomic_save_workbook()
# =============================================================================

class TestAtomicSaveWorkbook:
    """Tests for the atomic_save_workbook function with openpyxl."""
    
    def test_saves_workbook_successfully(self, tmp_path):
        """Should save an openpyxl workbook atomically."""
        # Arrange
        file_path = tmp_path / "test.xlsx"
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "Test Value"
        
        # Act
        atomic_save_workbook(wb, str(file_path))
        
        # Assert: File should exist with correct content
        assert file_path.exists()
        
        # Verify content by loading the saved file
        from openpyxl import load_workbook
        loaded_wb = load_workbook(str(file_path))
        assert loaded_wb.active["A1"].value == "Test Value"
    
    def test_preserves_original_on_save_error(self, tmp_path):
        """If save fails, original file should be preserved."""
        # Arrange: Create an existing file
        file_path = tmp_path / "existing.xlsx"
        original_wb = Workbook()
        original_wb.active["A1"] = "Original"
        original_wb.save(str(file_path))
        
        # Create a mock workbook that raises on save
        mock_wb = Mock()
        mock_wb.save.side_effect = Exception("Save failed")
        
        # Act & Assert
        with pytest.raises(Exception, match="Save failed"):
            atomic_save_workbook(mock_wb, str(file_path))
        
        # Original should be untouched
        from openpyxl import load_workbook
        loaded_wb = load_workbook(str(file_path))
        assert loaded_wb.active["A1"].value == "Original"
    
    def test_no_orphaned_temp_files_on_success(self, tmp_path):
        """After successful save, no temp files should remain."""
        # Arrange
        file_path = tmp_path / "test.xlsx"
        wb = Workbook()
        wb.active["A1"] = "Test"
        
        # Act
        atomic_save_workbook(wb, str(file_path))
        
        # Assert: No temp files in directory
        files = list(tmp_path.iterdir())
        temp_files = [f for f in files if f.name.startswith(TEMP_FILE_PREFIX)]
        assert len(temp_files) == 0
    
    def test_no_orphaned_temp_files_on_failure(self, tmp_path):
        """After failed save, temp files should be cleaned up."""
        # Arrange
        file_path = tmp_path / "test.xlsx"
        mock_wb = Mock()
        mock_wb.save.side_effect = Exception("Save failed")
        
        # Act
        try:
            atomic_save_workbook(mock_wb, str(file_path))
        except Exception:
            pass
        
        # Assert: No temp files in directory
        files = list(tmp_path.iterdir())
        temp_files = [f for f in files if f.name.startswith(TEMP_FILE_PREFIX)]
        assert len(temp_files) == 0


# =============================================================================
# INTEGRATION TEST: Full Workflow
# =============================================================================

class TestIntegration:
    """Integration tests for the complete atomic write workflow."""
    
    def test_modify_and_save_workbook(self, tmp_path):
        """
        End-to-end test: Load, modify, and atomically save a workbook.
        
        This simulates the actual use case in the application.
        """
        from openpyxl import load_workbook
        from openpyxl.styles import PatternFill
        
        # Arrange: Create initial workbook
        file_path = tmp_path / "data.xlsx"
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "Header"
        ws["A2"] = "Data Row 1"
        ws["A3"] = "Data Row 2"
        wb.save(str(file_path))
        
        # Act: Load, modify (highlight), and atomically save
        wb = load_workbook(str(file_path))
        ws = wb.active
        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        ws["A2"].fill = yellow_fill
        
        atomic_save_workbook(wb, str(file_path))
        
        # Assert: Verify modifications were saved
        loaded_wb = load_workbook(str(file_path))
        loaded_ws = loaded_wb.active
        assert loaded_ws["A1"].value == "Header"
        assert loaded_ws["A2"].value == "Data Row 1"
        assert loaded_ws["A2"].fill.start_color.rgb == "00FFFF00"  # Yellow
