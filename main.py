import pandas as pd
from openpyxl import load_workbook
import re

def getKeyColumns(codes_string):
    ans = {}
    codes_list = codes_string.split('\n')
    codes = [line.split()[1:] for line in codes_list]
    
    count = 1
    for row in codes:
        # Check if this row has a date (where the last one is '00:00:00')
        if row[-1] == '00:00:00':
            # Remove the time component
            row.pop()
    
    for i in codes:
        #empty string key
        key = ''
        #add all elements of i to key
        for j in i:
            key += j + '-'
        
        #add key to ans
        if key not in ans:
            ans[key] = count
            #print(count, ' : ', key)
        count += 1
    return ans

def getChangedRows(curr, prev):
    ans = []
    for i in curr:
        if i not in prev:
            ans.append([curr[i], i])
    return ans

def clear_filters(file_path):
    try:
        # Load the workbook
        wb = load_workbook(file_path)
        
        # Clear filters from each sheet
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            if ws.auto_filter:
                ws.auto_filter.ref = None
        
        # Save the workbook
        wb.save(file_path)
        print("Filters cleared successfully from all sheets")
    except Exception as e:
        print(f"Error clearing filters: {e}")

def read_excel_file(file_path, sheet_name=0):
    try:
        # Clear filters from all sheets
        clear_filters(file_path)
        
        # Get list of all sheet names
        xl = pd.ExcelFile(file_path)
        print("\nAvailable sheets:", xl.sheet_names)
        
        # Read the new and previous sheets
        current_sheet = len(xl.sheet_names)-2           # 2nd last sheet - New sheet
        previous_current_sheet = len(xl.sheet_names)-3    # 3rd last sheet - Previous sheet
        df_current = pd.read_excel(file_path, sheet_name=current_sheet)
        df_previous = pd.read_excel(file_path, sheet_name=previous_current_sheet)

        # Convert SAP columns to integers, handling any decimal values
        for df in [df_current, df_previous]:
            df['SAP ORDER'] = df['SAP ORDER'].fillna(0).astype(int)
            df['SAP CODE'] = df['SAP CODE'].fillna(0).astype(int)
        
        #Get key columns from new and previous sheets
        print(f"\nReading sheet: {xl.sheet_names[current_sheet]}")
        current_codes = getKeyColumns(df_current[['CUSTOMER CODE','SAP ORDER', 'SAP CODE', 'DESCRIPTION', 'PO', 'PRODUCED QTY', 'INVOICE', 'ESTIMATED DELIVERY DATE']].to_string())
        print('Found',len(current_codes),'rows in new sheet\n---')
        print(f"\nReading sheet: {xl.sheet_names[previous_current_sheet]}")
        previous_codes = getKeyColumns(df_previous[['CUSTOMER CODE', 'SAP ORDER', 'SAP CODE', 'DESCRIPTION', 'PO', 'PRODUCED QTY', 'INVOICE', 'ESTIMATED DELIVERY DATE']].to_string())
        print('Found',len(previous_codes),'rows in previous sheet\n---\n')

        #Get changed rows from new and previous sheets
        changed_rows = getChangedRows(current_codes, previous_codes)
        print('Found',len(changed_rows),'Significant Changes\n---')

         # Load workbook for highlighting
        wb = load_workbook(file_path)
        ws = wb[xl.sheet_names[current_sheet]]
        
        # Define highlight style (yellow background)
        from openpyxl.styles import PatternFill
        highlight_fill = PatternFill(start_color='FFFF00',
                                   end_color='FFFF00',
                                   fill_type='solid')
        
        # Highlight changed rows
        for row_info in changed_rows:
            row_num = row_info[0]  # Get the row number
            # Highlight all cells in the row
            for cell in ws[row_num]:
                cell.fill = highlight_fill
        
        # Save the workbook
        wb.save(file_path)
        print(f"Successfully highlighted {len(changed_rows)} changed rows")
    

        for i in changed_rows:
            print(i)

        return df_current

    except Exception as e:
        print(f"Error reading Excel file: {e}")
        return None

if __name__ == "__main__":
    excel_file_path = "test1.xlsx"
    # Read first sheet (by index)
    print("Reading first sheet:")
    df1 = read_excel_file(excel_file_path, sheet_name=0)
