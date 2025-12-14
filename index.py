import pandas as pd
import re
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# --- Configuration ---
IGNORED_COLUMNS = ['LOT #'] 
NOISE_THRESHOLD = 0.5 

def normalize_value(val):
    """Standardizes data for comparison."""
    if pd.isna(val):
        return "" 
    if isinstance(val, float) and val.is_integer():
        return str(int(val))
    if isinstance(val, (int, float)):
        return str(val)
    return str(val).strip()

def get_target_sheets(file_path):
    """Finds the last two sheets matching the pattern."""
    try:
        xls = pd.ExcelFile(file_path, engine='openpyxl')
        all_sheets = xls.sheet_names
        pattern = r"^\d{1,2}\.\d{1,2}"
        target_sheets = [name for name in all_sheets if re.match(pattern, name)]
        
        if len(target_sheets) < 2:
            print(f"Error: Need at least 2 matching sheets. Found: {len(target_sheets)}")
            return None, None
        return target_sheets[-2], target_sheets[-1]
    except Exception as e:
        print(f"Error finding sheets: {e}")
        return None, None

def get_common_columns_map(df_prev, df_curr):
    """Identifies common columns by Header Name."""
    def get_header_map(df):
        row_series = df.iloc[0] 
        value_map = {}
        for col_index, val in row_series.items():
            clean_val = normalize_value(val)
            if clean_val and clean_val not in IGNORED_COLUMNS: 
                if clean_val not in value_map:
                    value_map[clean_val] = []
                value_map[clean_val].append(col_index)
        return value_map

    map_prev = get_header_map(df_prev)
    map_curr = get_header_map(df_curr)
    common_values = set(map_prev.keys()) & set(map_curr.keys())
    
    results = {}
    for val in common_values:
        results[val] = {
            "prev_indices": sorted(map_prev[val]),
            "curr_indices": sorted(map_curr[val])
        }
    return results

def filter_noisy_columns(df_prev, df_curr, matches_data):
    """Removes columns if they change too much between sheets."""
    clean_matches = {}
    data_prev = df_prev.iloc[1:].reset_index(drop=True)
    data_curr = df_curr.iloc[1:].reset_index(drop=True)
    min_len = min(len(data_prev), len(data_curr))
    
    if min_len == 0:
        return matches_data 

    print("\n--- Checking for Noisy Columns ---")
    for col_name, indices in matches_data.items():
        idx_prev = indices['prev_indices'][0]
        idx_curr = indices['curr_indices'][0]
        col_vals_prev = data_prev.iloc[:min_len, idx_prev].apply(normalize_value).tolist()
        col_vals_curr = data_curr.iloc[:min_len, idx_curr].apply(normalize_value).tolist()

        diff_count = sum(1 for p, c in zip(col_vals_prev, col_vals_curr) if p != c)
        change_ratio = diff_count / min_len

        if change_ratio >= NOISE_THRESHOLD:
            print(f"Ignoring '{col_name}': {change_ratio:.0%} of rows changed")
        else:
            clean_matches[col_name] = indices
    return clean_matches

def generate_signatures(df, matches_data, sheet_key):
    """Generates signatures using the pre-loaded DataFrame."""
    sorted_keys = sorted(matches_data.keys())
    target_columns = []
    for key in sorted_keys:
        target_columns.extend(matches_data[key][sheet_key])

    signatures = []
    data_df = df.iloc[1:]
    
    for _, row in data_df.iterrows():
        row_values = []
        for col_idx in target_columns:
            if col_idx < len(row):
                val = normalize_value(row.iloc[col_idx])
                row_values.append(val)
            else:
                row_values.append("")
        signature = "-".join(row_values)
        signatures.append(signature)
    return signatures

def highlight_rows(file_path, sheet_name, indices):
    """
    Opens the Excel file and highlights the specified DataFrame indices in Yellow.
    """
    try:
        wb = load_workbook(file_path)
        if sheet_name not in wb.sheetnames:
            print(f"Error: Sheet '{sheet_name}' not found for highlighting.")
            return

        ws = wb[sheet_name]
        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

        print(f"\nHighlighting {len(indices)} rows in '{sheet_name}'...")

        # Calculate offset: 
        # skiprows=1 means Excel Row 1 is skipped.
        # DataFrame Index 0 is Excel Row 2 (Header).
        # DataFrame Index 1 is Excel Row 3 (Data).
        # Formula: Excel Row = DataFrame Index + 2
        for idx in indices:
            excel_row = idx + 2
            # Iterate through all cells in that row and color them
            for cell in ws[excel_row]:
                cell.fill = yellow_fill

        wb.save(file_path)
        print("File saved successfully with highlights.")
        
    except Exception as e:
        print(f"Error highlighting rows: {e}")

def get_changed_rows(filename):
    script_dir = os.path.dirname(os.path.abspath(__file__))
    file_path = os.path.join(script_dir, filename)

    if not os.path.exists(file_path):
        print("File not found.")
        return None

    prev_sheet_name, curr_sheet_name = get_target_sheets(file_path)
    if not prev_sheet_name:
        return None

    print(f"Processing: '{prev_sheet_name}' -> '{curr_sheet_name}'")

    df_prev = pd.read_excel(file_path, sheet_name=prev_sheet_name, header=None, skiprows=1)
    df_curr = pd.read_excel(file_path, sheet_name=curr_sheet_name, header=None, skiprows=1)

    matches_data = get_common_columns_map(df_prev, df_curr)
    if not matches_data:
        print("No common columns found.")
        return None

    matches_data = filter_noisy_columns(df_prev, df_curr, matches_data)
    if not matches_data:
        print("All columns were filtered out.")
        return None

    print("\nGenerating signatures...")
    sigs_prev = set(generate_signatures(df_prev, matches_data, 'prev_indices'))
    sigs_curr = generate_signatures(df_curr, matches_data, 'curr_indices')

    changed_indices = []
    for idx, sig in enumerate(sigs_curr):
        if sig not in sigs_prev:
            # +1 because sigs_curr starts from df.iloc[1]
            changed_indices.append(idx + 1) 

    if changed_indices:
        # Call the new highlighter function
        highlight_rows(file_path, curr_sheet_name, changed_indices)
        return df_curr.iloc[changed_indices]
    else:
        return pd.DataFrame()

# --- Execution ---
if __name__ == "__main__":
    file_name = 'data3.xlsx' 
    
    changed_rows = get_changed_rows(file_name)
    
    if changed_rows is not None and not changed_rows.empty:
        print(f"\nOperation complete. Check '{file_name}' for yellow highlights.")
    elif changed_rows is not None:
        print("\nNo changes detected.")